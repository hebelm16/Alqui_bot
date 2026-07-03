import io
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MESES_NOMBRES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def exportar_informe_excel(mes: int, anio: int, datos: dict) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) estructurado con el reporte financiero del mes.
    Contiene pestañas de Resumen, Pagos y Gastos.
    Devuelve un buffer io.BytesIO listo para ser descargado/enviado.
    """
    wb = openpyxl.Workbook()
    
    # Colores corporativos
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    accent_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_title = Font(name="Calibri", size=16, bold=True, color="1A365D")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )
    border_double = Border(
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='double', color='1A365D')
    )

    currency_format = '"RD$"#,##0.00'
    date_format = 'DD/MM/YYYY'

    nombre_mes = MESES_NOMBRES[mes] if 1 <= mes <= 12 else str(mes)

    # ----------------------------------------------------
    # Pestaña 1: Resumen General
    # ----------------------------------------------------
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    ws_resumen.views.sheetView[0].showGridLines = True

    ws_resumen["B2"] = f"INFORME FINANCIERO MENSUAL — {nombre_mes.upper()} {anio}"
    ws_resumen["B2"].font = font_title

    ws_resumen["B4"] = "Concepto Financiero"
    ws_resumen["C4"] = "Monto (RD$)"
    ws_resumen["B4"].font = font_header
    ws_resumen["C4"].font = font_header
    ws_resumen["B4"].fill = header_fill
    ws_resumen["C4"].fill = header_fill
    ws_resumen["B4"].alignment = Alignment(horizontal="center")
    ws_resumen["C4"].alignment = Alignment(horizontal="center")

    kpis = [
        ("Ingresos Totales (Pagos)", float(datos.get("total_ingresos", 0))),
        ("Gastos Totales Registrados", float(datos.get("total_gastos", 0))),
        ("Comisión Hecbel (5%)", float(datos.get("total_comision", 0))),
        ("Monto Neto a Entregar", float(datos.get("monto_neto", 0)))
    ]

    row_idx = 5
    for concepto, monto in kpis:
        cell_b = ws_resumen.cell(row=row_idx, column=2, value=concepto)
        cell_c = ws_resumen.cell(row=row_idx, column=3, value=monto)
        
        cell_b.font = font_bold if "Neto" in concepto else font_normal
        cell_c.font = font_bold if "Neto" in concepto else font_normal
        cell_c.number_format = currency_format
        
        cell_b.border = border_thin
        cell_c.border = border_thin
        
        if "Neto" in concepto:
            cell_b.fill = total_fill
            cell_c.fill = total_fill
            cell_c.border = border_double
        elif row_idx % 2 == 0:
            cell_b.fill = accent_fill
            cell_c.fill = accent_fill
            
        row_idx += 1

    # ----------------------------------------------------
    # Pestaña 2: Pagos del Mes
    # ----------------------------------------------------
    ws_pagos = wb.create_sheet(title="Pagos del Mes")
    ws_pagos.views.sheetView[0].showGridLines = True
    
    ws_pagos["A1"] = f"DETALLE DE INGRESOS — {nombre_mes.upper()} {anio}"
    ws_pagos["A1"].font = Font(name="Calibri", size=14, bold=True, color="1A365D")

    headers_pagos = ["ID", "Fecha de Pago", "Inquilino", "Monto (RD$)"]
    for col_idx, header in enumerate(headers_pagos, start=1):
        cell = ws_pagos.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center")

    pagos = datos.get("pagos_mes", [])
    row_idx = 4
    for pago in pagos:
        p_id, p_fecha, p_inq, p_monto = pago[0], pago[1], pago[2], float(pago[3])
        
        c_id = ws_pagos.cell(row=row_idx, column=1, value=p_id)
        c_fecha = ws_pagos.cell(row=row_idx, column=2, value=p_fecha if isinstance(p_fecha, (datetime, date)) else str(p_fecha))
        c_inq = ws_pagos.cell(row=row_idx, column=3, value=p_inq)
        c_monto = ws_pagos.cell(row=row_idx, column=4, value=p_monto)
        
        c_id.alignment = Alignment(horizontal="center")
        if isinstance(p_fecha, (datetime, date)):
            c_fecha.number_format = date_format
        c_monto.number_format = currency_format
        
        for col in range(1, 5):
            ws_pagos.cell(row=row_idx, column=col).border = border_thin
            if row_idx % 2 == 1:
                ws_pagos.cell(row=row_idx, column=col).fill = accent_fill
        row_idx += 1

    # Fila de Total en Pagos
    if pagos:
        ws_pagos.cell(row=row_idx, column=3, value="Total Ingresos").font = font_bold
        total_p = ws_pagos.cell(row=row_idx, column=4, value=f"=SUM(D4:D{row_idx-1})")
        total_p.font = font_bold
        total_p.number_format = currency_format
        total_p.border = border_double

    # ----------------------------------------------------
    # Pestaña 3: Gastos del Mes
    # ----------------------------------------------------
    ws_gastos = wb.create_sheet(title="Gastos del Mes")
    ws_gastos.views.sheetView[0].showGridLines = True
    
    ws_gastos["A1"] = f"DETALLE DE GASTOS — {nombre_mes.upper()} {anio}"
    ws_gastos["A1"].font = Font(name="Calibri", size=14, bold=True, color="1A365D")

    headers_gastos = ["ID", "Fecha del Gasto", "Descripción", "Monto (RD$)"]
    for col_idx, header in enumerate(headers_gastos, start=1):
        cell = ws_gastos.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center")

    gastos = datos.get("gastos_mes", [])
    row_idx = 4
    for gasto in gastos:
        g_id, g_fecha, g_desc, g_monto = gasto[0], gasto[1], gasto[2], float(gasto[3])
        
        c_id = ws_gastos.cell(row=row_idx, column=1, value=g_id)
        c_fecha = ws_gastos.cell(row=row_idx, column=2, value=g_fecha if isinstance(g_fecha, (datetime, date)) else str(g_fecha))
        c_desc = ws_gastos.cell(row=row_idx, column=3, value=g_desc)
        c_monto = ws_gastos.cell(row=row_idx, column=4, value=g_monto)
        
        c_id.alignment = Alignment(horizontal="center")
        if isinstance(g_fecha, (datetime, date)):
            c_fecha.number_format = date_format
        c_monto.number_format = currency_format
        
        for col in range(1, 5):
            ws_gastos.cell(row=row_idx, column=col).border = border_thin
            if row_idx % 2 == 1:
                ws_gastos.cell(row=row_idx, column=col).fill = accent_fill
        row_idx += 1

    if gastos:
        ws_gastos.cell(row=row_idx, column=3, value="Total Gastos").font = font_bold
        total_g = ws_gastos.cell(row=row_idx, column=4, value=f"=SUM(D4:D{row_idx-1})")
        total_g.font = font_bold
        total_g.number_format = currency_format
        total_g.border = border_double

    # Autoajustar anchos de columna en todas las hojas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
